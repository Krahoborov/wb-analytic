import logging
import asyncio
import openpyxl
from openpyxl import load_workbook
import io
from datetime import datetime, timedelta
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, InputFile
from tg_bot.models import sessionmaker, engine
from tg_bot.models import Shop, CashedShopData, Advertisement, Penalty
from tg_bot.models import (
    TaxSystemType, TaxSystemSetting,
    ProductCost, RegularExpense, OneTimeExpense
)
from tg_bot.keyboards.pnl_menu import pnl_period_keyboard
from tg_bot.services.wb_api import fetch_full_report, fetch_report_detail_by_period
from tg_bot.states.pnl_states import PNLStates
from dateutil.relativedelta import relativedelta

logger = logging.getLogger(__name__)

# Главное меню P&L
async def pnl_callback(callback: types.CallbackQuery, state: FSMContext):
    # Проверяем выбран ли магазин
    async with state.proxy() as data:
        if 'shop' not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return
    
    keyboard = pnl_period_keyboard()
    await callback.message.answer(
        "📊 <b>Расчёт прибыли и убытков (P&L)</b>\n\n"
        "Выберите период для расчета:",
        reply_markup=keyboard
    )
    await PNLStates.waiting_for_period.set()

async def generate_pnl_excel_report(shop_id: int, shop_api_token: str, start_date: datetime, end_date: datetime, shop_name: str):
    """Генерация Excel-отчета PNL"""


    session = sessionmaker()(bind=engine)
    # Получаем свежие данные из WB API

    if True:
        loop = asyncio.get_event_loop()
        report_data = await loop.run_in_executor(
            None,
            fetch_report_detail_by_period,
            shop_api_token,
            start_date,
            end_date
        )    
    #report_data = await fetch_report_detail_by_period(shop_api_token, start_date, end_date)
    if not report_data:
        return None
    try:
        # Получаем кэшированные данные
        cashed_data = (
            session.query(CashedShopData)
            .filter(CashedShopData.shop_id == shop_id)
            .first()
        )
        if not cashed_data or not cashed_data.cashed_all:
            return None


        # --- ДОБАВЛЕНО: Годовой отчёт ---
        is_year_report = (
            start_date.month == 1 and start_date.day == 1 and
            (end_date.year == start_date.year)
        )

        if is_year_report:
            try:
                wb = load_workbook("pnl_template_year.xlsx")
            except FileNotFoundError:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "P&L Year"
            ws = wb.active

            year = start_date.year
            # Период текущего года: с 1 января по end_date
            year_start = datetime(year, 1, 1)
            year_end = end_date
            # Период прошлого года: с 1 января по 31 декабря прошлого года
            prev_year_start = datetime(year - 1, 1, 1)
            prev_year_end = datetime(year - 1, 12, 31)

            # Данные за текущий и прошлый год
            current_year_data = []
            prev_year_data = []
            for item in report_data:
                try:
                    sale_date = datetime.strptime(item.get("sale_dt", "")[:10], "%Y-%m-%d")
                    if year_start <= sale_date <= year_end:
                        current_year_data.append(item)
                    elif prev_year_start <= sale_date <= prev_year_end:
                        prev_year_data.append(item)
                except (ValueError, TypeError):
                    continue

            current_metrics = await calculate_metrics_from_report(
                current_year_data, shop_id, year_start, year_end
            )
            previous_metrics = await calculate_metrics_from_report(
                prev_year_data, shop_id, prev_year_start, prev_year_end
            )

            current_values = [
                current_metrics.get("orders", 0),
                current_metrics.get("sales", 0),
                current_metrics.get("commission", 0),
                current_metrics.get("cost_of_goods", 0),
                current_metrics.get("tax", 0),
                current_metrics.get("logistics", 0),
                current_metrics.get("storage", 0),
                current_metrics.get("stops", 0),
                current_metrics.get("advert", 0),
                current_metrics.get("revenue", 0),
                current_metrics.get("net_profit", 0)
            ]
            previous_values = [
                previous_metrics.get("orders", 0),
                previous_metrics.get("sales", 0),
                previous_metrics.get("commission", 0),
                previous_metrics.get("cost_of_goods", 0),
                previous_metrics.get("tax", 0),
                previous_metrics.get("logistics", 0),
                previous_metrics.get("storage", 0),
                previous_metrics.get("stops", 0),
                previous_metrics.get("advert", 0),
                previous_metrics.get("revenue", 0),
                previous_metrics.get("net_profit", 0)
            ]

            # C3–C13: метрики за год
            for i, value in enumerate(current_values, 3):
                ws[f'C{i}'] = value

            # E3–E13: динамика год к году
            for i in range(3, 14):
                curr = current_values[i-3]
                prev = previous_values[i-3]
                if prev != 0:
                    ws[f'E{i}'] = (curr - prev) / prev
                else:
                    ws[f'E{i}'] = 0 if curr == 0 else 1

            # F3–Q13: метрики по месяцам (F — январь, ..., Q — декабрь)
            for month in range(1, 13):
                month_start = datetime(year, month, 1)
                if month == 12:
                    month_end = datetime(year, 12, 31)
                else:
                    month_end = datetime(year, month + 1, 1) - timedelta(days=1)
                if month_end > end_date:
                    month_end = end_date
                if month_start > end_date:
                    continue
                month_data = [
                    item for item in current_year_data
                    if month_start <= datetime.strptime(item.get("sale_dt", "")[:10], "%Y-%m-%d") <= month_end
                ]
                if not month_data:
                    continue  # Если данных нет, ячейки остаются пустыми
                month_metrics = await calculate_metrics_from_report(month_data, shop_id, month_start, month_end)
                month_values = [
                    month_metrics.get("orders", 0),
                    month_metrics.get("sales", 0),
                    month_metrics.get("commission", 0),
                    month_metrics.get("cost_of_goods", 0),
                    month_metrics.get("tax", 0),
                    month_metrics.get("logistics", 0),
                    month_metrics.get("storage", 0),
                    month_metrics.get("stops", 0),
                    month_metrics.get("advert", 0),
                    month_metrics.get("revenue", 0),
                    month_metrics.get("net_profit", 0)
                ]
                col_letter = openpyxl.utils.get_column_letter(6 + month - 1)  # F=6, ..., Q=17
                for row, value in enumerate(month_values, 3):
                    ws[f'{col_letter}{row}'] = value

            return wb

        # --- ДОБАВЛЕНО: Недельный отчёт ---
        is_week_report = (
            (end_date - start_date).days <= 8 and  # 7 или 8 дней (на всякий случай)
            start_date.weekday() == 0  # начинается с понедельника
        )


        if is_week_report:
            try:
                wb = load_workbook("pnl_template.xlsx")
            except FileNotFoundError:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "P&L Отчет"
            ws = wb.active

            # Метрики за текущую и прошлую неделю
            week_start = start_date
            week_end = end_date
            prev_week_start = week_start - timedelta(days=7)
            prev_week_end = week_end - timedelta(days=7)

            # Данные за текущую и прошлую неделю
            current_week_data = []
            prev_week_data = []
            for item in report_data:
                try:
                    sale_date = datetime.strptime(item.get("sale_dt", "")[:10], "%Y-%m-%d")
                    if week_start <= sale_date <= week_end:
                        current_week_data.append(item)
                    elif prev_week_start <= sale_date <= prev_week_end:
                        prev_week_data.append(item)
                except (ValueError, TypeError):
                    continue

            current_metrics = await calculate_metrics_from_report(
                current_week_data, shop_id, week_start, week_end
            )
            previous_metrics = await calculate_metrics_from_report(
                prev_week_data, shop_id, prev_week_start, prev_week_end
            )

            current_values = [
                current_metrics.get("orders", 0),
                current_metrics.get("sales", 0),
                current_metrics.get("commission", 0),
                current_metrics.get("cost_of_goods", 0),
                current_metrics.get("tax", 0),
                current_metrics.get("logistics", 0),
                current_metrics.get("storage", 0),
                current_metrics.get("stops", 0),
                current_metrics.get("advert", 0),
                current_metrics.get("revenue", 0),
                current_metrics.get("net_profit", 0)
            ]
            previous_values = [
                previous_metrics.get("orders", 0),
                previous_metrics.get("sales", 0),
                previous_metrics.get("commission", 0),
                previous_metrics.get("cost_of_goods", 0),
                previous_metrics.get("tax", 0),
                previous_metrics.get("logistics", 0),
                previous_metrics.get("storage", 0),
                previous_metrics.get("stops", 0),
                previous_metrics.get("advert", 0),
                previous_metrics.get("revenue", 0),
                previous_metrics.get("net_profit", 0)
            ]

            # C3–C13: метрики за неделю
            for i, value in enumerate(current_values, 3):
                ws[f'C{i}'] = value

            # E3–E13: динамика неделя к неделе
            for i in range(3, 14):
                curr = current_values[i-3]
                prev = previous_values[i-3]
                if prev != 0:
                    ws[f'E{i}'] = (curr - prev) / prev
                else:
                    ws[f'E{i}'] = 0 if curr == 0 else 1

            # F3–L13: метрики по дням недели (F=понедельник, ..., L=воскресенье)
            for day_offset in range(7):
                day_date = week_start + timedelta(days=day_offset)
                if day_date > week_end:
                    break
                day_data = [
                    item for item in current_week_data
                    if datetime.strptime(item.get("sale_dt", "")[:10], "%Y-%m-%d").date() == day_date.date()
                ]
                if not day_data:
                    continue  # Если данных нет, ячейки остаются пустыми
                day_metrics = await calculate_metrics_from_report(day_data, shop_id, day_date, day_date)
                day_values = [
                    day_metrics.get("orders", 0),
                    day_metrics.get("sales", 0),
                    day_metrics.get("commission", 0),
                    day_metrics.get("cost_of_goods", 0),
                    day_metrics.get("tax", 0),
                    day_metrics.get("logistics", 0),
                    day_metrics.get("storage", 0),
                    day_metrics.get("stops", 0),
                    day_metrics.get("advert", 0),
                    day_metrics.get("revenue", 0),
                    day_metrics.get("net_profit", 0)
                ]
                col_letter = openpyxl.utils.get_column_letter(6 + day_offset)  # F=6, ..., L=12
                for row, value in enumerate(day_values, 3):
                    ws[f'{col_letter}{row}'] = value
            print("week report is active")
            return wb



        # Рассчитываем периоды для сравнения
        period_days = (end_date - start_date).days
        
        # Период для сравнения (предыдущий такой же длины)
        previous_start = start_date - timedelta(days=period_days)
        previous_end = start_date - timedelta(days=1)

        # Фильтруем данные за текущий период
        current_report_data = []
        for item in report_data:
            try:
                sale_date = datetime.strptime(item.get("sale_dt", "")[:10], "%Y-%m-%d")
                if start_date <= sale_date <= end_date:
                    current_report_data.append(item)
            except (ValueError, TypeError):
                continue

        # Фильтруем данные за предыдущий период (для расчета динамики)
        previous_report_data = []
        for item in cashed_data.cashed_all:
            try:
                sale_date = datetime.strptime(item.get("sale_dt", "")[:10], "%Y-%m-%d")
                if previous_start <= sale_date <= previous_end:
                    previous_report_data.append(item)
            except (ValueError, TypeError):
                continue

        if not current_report_data:
            return None

        # Загружаем шаблон
        try:
            wb = load_workbook("pnl_template.xlsx")
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "P&L Отчет"
        
        ws = wb.active

        # Рассчитываем метрики за текущий месяц
        current_metrics = await calculate_metrics_from_report(current_report_data, shop_id, start_date, end_date)
        
        # Рассчитываем метрики за предыдущий период (только для расчета динамики)
        previous_metrics = await calculate_metrics_from_report(previous_report_data, shop_id, previous_start, previous_end)
        
        # Заполняем значения в ячейки C3-C13 (текущий период)
        current_values = [
            current_metrics.get("orders", 0),           # C3 - Заказы
            current_metrics.get("sales", 0),            # C4 - Выкупы
            current_metrics.get("commission", 0),       # C5 - Комиссия
            current_metrics.get("cost_of_goods", 0),    # C6 - Себестоимость
            current_metrics.get("tax", 0),              # C7 - Налог
            current_metrics.get("logistics", 0),        # C8 - Логистика
            current_metrics.get("storage", 0),          # C9 - Хранение
            current_metrics.get("stops", 0),            # C10 - Штрафы и корректировки
            current_metrics.get("advert", 0),           # C11 - Реклама
            current_metrics.get("revenue", 0),          # C12 - К перечислению (выручка)
            current_metrics.get("net_profit", 0)        # C13 - Чистая прибыль
        ]

        # Значения за предыдущий период (только для расчета динамики)
        previous_values = [
            previous_metrics.get("orders", 0),           # Заказы
            previous_metrics.get("sales", 0),            # Выкупы  
            previous_metrics.get("commission", 0),       # Комиссия
            previous_metrics.get("cost_of_goods", 0),    # Себестоимость
            previous_metrics.get("tax", 0),              # Налог
            previous_metrics.get("logistics", 0),        # Логистика
            previous_metrics.get("storage", 0),          # Хранение
            previous_metrics.get("stops", 0),            # Штрафы и корректировки
            previous_metrics.get("advert", 0),           # Реклама
            previous_metrics.get("revenue", 0),          # К перечислению (выручка)
            previous_metrics.get("net_profit", 0)        # Чистая прибыль
        ]

        # Заполняем ячейки C3-C13 (текущий период)
        for i, value in enumerate(current_values, 3):
            ws[f'C{i}'] = value

            # Рассчитываем и заполняем динамику в ячейки E3-E13
            # Формула: (текущий - предыдущий) / предыдущий
            for i in range(3, 14):
                current_val = current_values[i-3]
                previous_val = previous_values[i-3]
                
                if previous_val != 0:
                    dynamic = (current_val - previous_val) / previous_val
                    ws[f'E{i}'] = dynamic
                else:
                    # Если предыдущее значение равно 0, устанавливаем 0 или 1
                    ws[f'E{i}'] = 0 if current_val == 0 else 1

            # ВОССТАНАВЛИВАЕМ: Заполняем ячейки C3-C13 (текущий период)
            for i, value in enumerate(current_values, 3):
                ws[f'C{i}'] = value

            # ДОБАВЛЯЕМ: Заполняем данные по дням месяца (F3-F13, G3-G13, и т.д.)
            # Собираем уникальные даты с данными
            # Проходим по всем дням периода, даже если нет продаж
            sorted_dates = []
            current_date = start_date.date()
            while current_date <= end_date.date():
                sorted_dates.append(current_date)
                current_date += timedelta(days=1)

            # Функция для получения буквы столбца по номеру
            def get_column_letter(column_number):
                """Преобразует номер столбца в букву (A=1, B=2, ..., Z=26, AA=27, AB=28, ...)"""
                result = ""
                while column_number > 0:
                    column_number -= 1
                    result = chr(65 + (column_number % 26)) + result
                    column_number //= 26
                return result

            # Заполняем данные только для дней с данными
            for day_index, current_date in enumerate(sorted_dates):
                # Фильтруем данные за конкретный день
                daily_data = []
                for item in cashed_data.cashed_all:
                    try:
                        sale_date = datetime.strptime(item.get("sale_dt", "")[:10], "%Y-%m-%d")
                        if sale_date.date() == current_date:
                            daily_data.append(item)
                    except (ValueError, TypeError):
                        continue
                
                # Рассчитываем метрики за день
                daily_metrics = await calculate_metrics_from_report(daily_data, shop_id, current_date, current_date)
                
                # Определяем столбец (F=6, G=7, H=8, и т.д.)
                column_number = 6 + day_index  # F=6, G=7, H=8, ..., Z=26, AA=27, AB=28, ...
                column_letter = get_column_letter(column_number)
                
                # Заполняем значения в ячейки F3-F13, G3-G13, и т.д.
                daily_values = [
                    daily_metrics.get("orders", 0),           # Заказы
                    daily_metrics.get("sales", 0),            # Выкупы  
                    daily_metrics.get("commission", 0),       # Комиссия
                    daily_metrics.get("cost_of_goods", 0),    # Себестоимость
                    daily_metrics.get("tax", 0),              # Налог
                    daily_metrics.get("logistics", 0),        # Логистика
                    daily_metrics.get("storage", 0),          # Хранение
                    daily_metrics.get("stops", 0),            # Штрафы и корректировки
                    daily_metrics.get("advert", 0),           # Реклама
                    daily_metrics.get("revenue", 0),          # К перечислению (выручка)
                    daily_metrics.get("net_profit", 0)        # Чистая прибыль
                ]
                
                for i, value in enumerate(daily_values, 3):
                    ws[f'{column_letter}{i}'] = value

            return wb
        
    except Exception as e:
        logger.error(f"Ошибка генерации PNL отчета: {e}")
        return None
    finally:
        session.close()

# Расчет показателей на основе отчета
# Обновляем функцию calculate_metrics_from_report для добавления недостающих метрик
async def calculate_metrics_from_report(report_data, shop_id, start_date, end_date):
    session = sessionmaker()(bind=engine)
    try:
        # Основные показатели
        revenue = 0
        logistics = 0
        storage_fee = 0
        commission = 0
        cost_of_goods = 0
        orders = 0
        sales = 0
        
        # Собираем артикулы для расчета себестоимости
        articles = {}
        for item in report_data:
            # Выручка
            retail_price = item.get('retail_price_withdisc_rub', 0)
            quantity = item.get('quantity', 0)
            revenue += retail_price * quantity
            
            # Заказы и продажи
            doc_type = item.get("doc_type_name", "")
            if "продажа" in doc_type.lower() or "sale" in doc_type.lower():
                sales += quantity
            orders += quantity
            
            # Логистика
            logistics += item.get('delivery_rub', 0)
            
            # Хранение
            storage_fee += item.get('storage_fee', 0)
            
            # Комиссия WB
            commission += item.get('ppvz_sales_commission', 0) + item.get('ppvz_vw', 0) + item.get('ppvz_vw_nds', 0)
            
            # Собираем данные для себестоимости
            article = item.get('nm_id')
            if article:
                if article not in articles:
                    articles[article] = 0
                articles[article] += quantity
        
        # Себестоимость
        for article, quantity in articles.items():
            product_cost = session.query(ProductCost).filter(
                ProductCost.shop_id == shop_id,
                ProductCost.article == article
            ).first()
            if product_cost:
                cost_of_goods += product_cost.cost * quantity
        
        # Налоговая ставка с поддержкой кастомного процента
        tax_setting = session.query(TaxSystemSetting).filter(
            TaxSystemSetting.shop_id == shop_id
        ).first()
        
        if tax_setting:
            if tax_setting.tax_system == TaxSystemType.USN_6:
                tax_rate = 0.06
                print("tax_rate = ", tax_rate)
            elif tax_setting.tax_system == TaxSystemType.NO_TAX:
                tax_rate = 0.0
                print("tax_rate = ", tax_rate)
            elif tax_setting.tax_system == TaxSystemType.CUSTOM:
                tax_rate = tax_setting.custom_percent / 100 #if tax_setting.custom_percent else 0.0
                print("tax_rate = ", tax_rate)
            else:
                tax_rate = 0.0
                print("пошло в else")
        else:
            tax_rate = 0.0
            
        tax = revenue * tax_rate
        
        # Регулярные затраты за период
        regular_expenses = 0
        days_in_period = (end_date - start_date).days + 1
        for expense in session.query(RegularExpense).filter(RegularExpense.shop_id == shop_id):
            if expense.frequency == "daily":
                regular_expenses += expense.amount * days_in_period
            elif expense.frequency == "weekly":
                regular_expenses += expense.amount * (days_in_period / 7)
            elif expense.frequency == "monthly":
                regular_expenses += expense.amount * (days_in_period / 30)

        # Рекламные затраты за период
        advert = sum(
            i.amount for i in session.query(Advertisement)
            .filter(Advertisement.shop_id == shop_id)
            .filter(Advertisement.date >= start_date)
            .filter(Advertisement.date <= end_date)
            .all()
        )

        # Штрафы за период
        stops = sum(
            i.sum for i in session.query(Penalty)
            .filter(Penalty.shop_id == shop_id)
            .filter(Penalty.date >= start_date)
            .filter(Penalty.date <= end_date)
            .all()
        )

        # Чистая прибыль
        net_profit = revenue - (commission + logistics + storage_fee + tax + cost_of_goods + regular_expenses + advert + stops)
        
        # Рентабельность (добавляем эту строку)
        profitability = (net_profit / revenue) * 100 if revenue > 0 else 0
        
        # Разовые затраты (инвестиционные)
        one_time_expenses = session.query(OneTimeExpense).filter(OneTimeExpense.shop_id == shop_id).all()
        total_one_time = sum(expense.amount for expense in one_time_expenses)
        
        # Срок окупаемости
        payback_period = "не определен"
        if net_profit > 0 and total_one_time > 0:
            months = total_one_time / net_profit
            payback_period = f"{months:.1f} месяцев"
        
        # ROI
        roi = "не определен"
        if total_one_time > 0:
            roi_value = (net_profit / total_one_time) * 100
            roi = f"{roi_value:.1f}%"
            if roi_value > 100:
                roi += " ✅ Поздравляем, вы окупили вложения!"
        
        return {
            "revenue": revenue,
            "commission": commission,
            "logistics": logistics,
            "storage": storage_fee,
            "cost_of_goods": cost_of_goods,
            "tax": tax,
            "regular_expenses": regular_expenses,
            "net_profit": net_profit,
            "profitability": profitability,  # ← ДОБАВЛЯЕМ ЭТО
            "payback_period": payback_period,  # ← И ЭТО
            "roi": roi,  # ← И ЭТО
            "advert": advert,
            "stops": stops,
            "orders": orders,
            "sales": sales
        }
    finally:
        session.close()
        session.close()

# Обработка выбора периода
# Обработка выбора периода
async def select_pnl_period_callback(callback: types.CallbackQuery, state: FSMContext):
    period_type = callback.data.split('_')[1]  # day, week, month, year
    await callback.message.edit_text(text="Подождите около 10 секунд, пока произведем подсчёт данных... (иногда дольше, но не более 2х минут)")
    
    # Определяем периоды
    now = datetime.utcnow()
    if period_type == "week":
        # Находим последний понедельник
        start_week = now - timedelta(days=now.isoweekday() - 1)
        start_date = datetime(start_week.year, start_week.month, start_week.day)
        end_date = now
        #start_date = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        period_name = "неделю"
    elif period_type == "month":
        #start_date = now - relativedelta(months=1)
        start_date = datetime(now.year, now.month, 1)
        end_date = now
        period_name = "месяц"
    elif period_type == "year":
        now = datetime.utcnow()
        #start_date = datetime(now.year, 1, 1)
        start_date = datetime(now.year, 1, 1)
        end_date = now
        period_name = "год"
    else:  # year
        start_date = now - relativedelta(years=1)
        end_date = now
        period_name = "год"
    
    async with state.proxy() as data:
        shop_id = data['shop']['id']
        shop_name = data['shop']['name'] or f"Магазин {shop_id}"
    
    # Показываем сообщение о загрузке
    await callback.message.edit_text(
        f"📊 <b>Генерация PNL отчета</b>\n\n"
        f"Магазин: {shop_name}\n"
        f"Период: за {period_name}\n\n"
        "Подождите, идет сбор и обработка данных..."
    )
    
    # В обработчике:
    async with state.proxy() as data:
        shop_id = data['shop']['id']
        shop_name = data['shop']['name'] or f"Магазин {shop_id}"
        shop_api_token = data['shop']['api_token']


    
    # Генерируем Excel отчет
    wb = await generate_pnl_excel_report(shop_id, shop_api_token, start_date, end_date, shop_name)
    
    if not wb:
        await callback.message.edit_text(
            "❌ <b>Не удалось сгенерировать отчет</b>\n\n"
            "Возможные причины:\n"
            "1. Нет данных за выбранный период\n"
            "2. Проблемы с подключением к базе данных\n"
            "3. Файл шаблона pnl_template.xlsx не найден"
        )
        return
    
    # Сохраняем в буфер
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    # Формируем имя файла
    safe_shop_name = "".join(c for c in shop_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    filename = f"pnl_{safe_shop_name}_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    
    # Отправляем файл
    file = InputFile(file_stream, filename=filename)
    await callback.message.answer_document(
        file,
        caption=f"📊 PNL отчет за {period_name}\nМагазин: {shop_name}"
    )
    
    # Возвращаемся к меню PNL
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("🔙 Назад", callback_data="pnl"))
    await callback.message.answer("✅ Отчет готов!", reply_markup=keyboard)

# Добавляем обработчик для Excel отчета
async def pnl_excel_callback(callback: types.CallbackQuery, state: FSMContext):
    # Определяем периоды
    now = datetime.utcnow()
    if period_type == "week":
        start_week = now - timedelta(days=now.isoweekday() - 1)
        start_date = datetime(start_week.year, start_week.month, start_week.day)

        #start_date = now - timedelta(weeks=1)
        end_date = now
        period_name = "неделю"
    elif period_type == "month":
        # С 1 числа текущего месяца
        #start_date = now.replace(day=1)  # ← ИЗМЕНИТЬ НА ЭТО
        start_date = datetime(now.year, now.month, 1)
        end_date = now
        period_name = "месяц"
    else:  # year
        #start_date = now - relativedelta(years=1)
        start_date = datetime(now.year, 1, 1)
        end_date = now
        period_name = "год"
    
    async with state.proxy() as data:
        shop_id = data['shop']['id']
        shop_name = data['shop']['name'] or f"Магазин {shop_id}"
    
    # Показываем сообщение о загрузке
    await callback.message.edit_text(
        f"📊 <b>Генерация PNL отчета</b>\n\n"
        f"Магазин: {shop_name}\n"
        f"Период: за {period_name}\n\n"
        "Подождите, идет сбор и обработка данных..."
    )
    
    # Генерируем Excel отчет
    wb = await generate_pnl_excel_report(shop_id, start_date, end_date, shop_name)
    
    if not wb:
        await callback.message.edit_text(
            "❌ <b>Не удалось сгенерировать отчет</b>\n\n"
            "Возможные причины:\n"
            "1. Нет данных за выбранный период\n"
            "2. Проблемы с подключением к базе данных\n"
            "3. Файл шаблона pnl_template.xlsx не найден"
        )
        return
    
    # Сохраняем в буфер
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    # Формируем имя файла
    safe_shop_name = "".join(c for c in shop_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    filename = f"pnl_{safe_shop_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    
    # Отправляем файл
    file = InputFile(file_stream, filename=filename)
    await callback.message.answer_document(
        file,
        caption=f"📊 PNL отчет за {period_name}\nМагазин: {shop_name}"
    )


def register_pnl_handlers(dp):
    dp.register_callback_query_handler(pnl_callback, text="pnl", state="*")
    dp.register_callback_query_handler(
        select_pnl_period_callback, 
        lambda c: c.data.startswith("pnlperiod_"), 
        state=PNLStates.waiting_for_period
    )
    dp.register_callback_query_handler(
        pnl_excel_callback,
        lambda c: c.data.startswith("pnl_excel_"),
        state="*"
    )
