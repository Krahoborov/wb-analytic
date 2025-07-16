from aiogram import types
from aiogram.dispatcher import FSMContext, Dispatcher
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, InputFile
from aiogram.utils.exceptions import MessageNotModified
from tg_bot.models import (
    Order,
    sessionmaker,
    engine,
    ProductCost,
    TaxSystemSetting,
    RegularExpense,
    TaxSystemType,
    CashedShopData,
    OneTimeExpense,
    Advertisement,
    Penalty,
User,
    RegularExpenseFrequency
)
from tg_bot.states.analytics_states import AnalyticsStates
from tg_bot.keyboards.analytics_menu import (
    analytics_menu_keyboard,
    period_keyboard,
    period_keyboard2,
)
from tg_bot.services.wb_api import fetch_full_report
from dateutil.relativedelta import relativedelta
from datetime import datetime, timedelta
import math
import io
import logging
import openpyxl
import json
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# Главное меню аналитики
async def analytics_callback(callback: types.CallbackQuery):
    text = (
        "📈 <b>Аналитика и рекомендации</b>\n\n"
        "Здесь вы можете получить детальную аналитику по вашему бизнесу на Wildberries. "
        "Выберите интересующий раздел:"
    )
    keyboard = analytics_menu_keyboard()
    await callback.message.edit_text(text, reply_markup=keyboard)


# Обработчики для подменю
PROFITABILITY_LEVELS = [
    {
        "min": -float("inf"),
        "max": 20,
        "name": "⚠️ Низкая доходность",
        "characteristics": "Плохая рентабельность, высокие риски или низкая маржинальность.",
        "reasons": "Высокая конкуренция, большие расходы на логистику/хранение, низкие наценки.",
        "conclusion": "Такой бизнес невыгоден, нужно пересматривать модель.",
        "recommendations": [
            "Срочно пересмотрите ценовую политику и себестоимость.",
            "Ищите более выгодных поставщиков или сокращайте логистические издержки.",
            "Проверьте скрытые расходы (хранение, возвраты, реклама) и оптимизируйте их.",
            "Если рост невозможен – рассмотрите закрытие или смену ниши.",
        ],
        "action": "Оптимизировать или уходить",
    },
    {
        "min": 20,
        "max": 40,
        "name": "⚠️ Ниже среднего",
        "characteristics": "Минимально приемлемая рентабельность, но требует оптимизации.",
        "reasons": "Средняя конкуренция, умеренные издержки.",
        "conclusion": "Высокий риск уйти в ноль или минус из-за внешних факторов.",
        "recommendations": [
            "Увеличивайте маржу через улучшение упаковки, допродажи или брендинг.",
            "Автоматизируйте процессы для снижения операционных затрат.",
            "Тестируйте новые рекламные каналы для увеличения продаж.",
            "Анализируйте конкурентов на предмет более выгодных товаров.",
        ],
        "action": "Улучшать и тестировать другие товары",
    },
    {
        "min": 40,
        "max": 60,
        "name": "✅ Средняя доходность",
        "characteristics": "Нормальный уровень для стабильного бизнеса.",
        "reasons": "Хороший спрос, грамотное ценообразование, контроль затрат.",
        "conclusion": "Устойчивый бизнес, можно масштабировать.",
        "recommendations": [
            "Фокусируйтесь на стабильности: контролируйте качество и сервис.",
            "Расширяйте ассортимент в нише для увеличения среднего чека.",
            "Инвестируйте в лояльность клиентов (отзывы, рассылки).",
            "Тестируйте смежные ниши с более высокой маржой.",
        ],
        "action": "Закрепляться и расти",
    },
    {
        "min": 60,
        "max": 100,
        "name": "🔥 Высокая доходность",
        "characteristics": "Очень хорошая рентабельность, перспективный бизнес.",
        "reasons": "Уникальный товар, низкая конкуренция, эффективные рекламные каналы.",
        "conclusion": "Отличный результат, стоит вкладывать больше ресурсов.",
        "recommendations": [
            "Активно масштабируйте: выходите на новые маркетплейсы или рынки.",
            "Усиливайте бренд и работайте с повторными продажами.",
            "Диверсифицируйте поставщиков для снижения рисков.",
            "Инвестируйте часть прибыли в новые высокомаржинальные товары.",
        ],
        "action": "Масштабировать и защищать",
    },
    {
        "min": 100,
        "max": float("inf"),
        "name": "✨ Премиальная доходность",
        "characteristics": "Высокомаржинальный бизнес, часто нишевый.",
        "reasons": "Эксклюзивные товары, VIP-сегмент, отсутствие прямых аналогов.",
        "conclusion": "Редкий и ценный кейс, требует защиты позиций.",
        "recommendations": [
            "Укрепляйте эксклюзивность через товарный знак и уникальные условия с поставщиками.",
            "Создавайте финансовую подушку безопасности.",
            "Масштабируйте до точки максимальной эффективности.",
            "Мониторьте динамику прибыли и будьте готовы к поиску новых товаров.",
        ],
        "action": "Укреплять позиции или выжимать все соки",
    },
]


def get_profitability_level(profitability):
    """Определение уровня доходности по проценту рентабельности"""
    for level in PROFITABILITY_LEVELS:
        if level["min"] <= profitability < level["max"]:
            return level
    return PROFITABILITY_LEVELS[0]  # По умолчанию низкая доходность


async def profitability_estimation_callback(
    callback: types.CallbackQuery, state: FSMContext
):
    """Обработчик оценки доходности"""
    # Проверяем выбран ли магазин
    async with state.proxy() as data:
        if "shop" not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return

    # Сохраняем контекст для пагинации
    async with state.proxy() as data:
        data["analytics_type"] = "profitability"
        data["article_page"] = 0

    await show_articles_page(callback, state)


async def calculate_profitability_for_article(article, shop_id, api_token):
    """Расчет доходности для конкретного артикула"""
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=30)

    # Получаем отчет за последний месяц
    report = await fetch_full_report(api_token, start_date, end_date)
    if not report:
        return None

    print("FLAG0")
    # for i in report:
    # print(i['nm_id'])
    # Фильтруем данные по выбранному артикулу
    article_data = [item for item in report if item.get("nm_id") == article]
    print("FLAG0.5")
    print(article_data)
    if not article_data:
        return None
    print("FLAG1")
    # Рассчитываем показатели
    total_quantity = sum(item.get("quantity", 0) for item in article_data)
    total_revenue = sum(
        item.get("retail_price_withdisc_rub", 0) for item in article_data
    )
    total_commission = -sum(
        item.get("ppvz_sales_commission", 0)
        + item.get("ppvz_vw", 0)
        + item.get("ppvz_vw_nds", 0)
        for item in article_data
    )
    print("FLAG2")

    # Рассчитываем общие расходы для распределения
    total_logistics = sum(item.get("delivery_rub", 0) for item in report)
    total_storage = sum(item.get("storage_fee", 0) for item in report)
    total_revenue_all = sum(item.get("retail_price_withdisc_rub", 0) for item in report)

    # Распределяем логистику и хранение пропорционально выручке
    logistics_share = (
        total_logistics * (total_revenue / total_revenue_all)
        if total_revenue_all
        else 0
    )
    storage_share = (
        total_storage * (total_revenue / total_revenue_all) if total_revenue_all else 0
    )
    print("FLAG3")
    # Получаем себестоимость
    session = sessionmaker()(bind=engine)
    try:
        product_cost = (
            session.query(ProductCost)
            .filter(ProductCost.shop_id == shop_id, ProductCost.article == article)
            .first()
        )
        cost_per_item = product_cost.cost if product_cost else 0
    finally:
        session.close()

    total_cost = cost_per_item * total_quantity
    total_expenses = total_commission + logistics_share + storage_share + total_cost

    # Рассчитываем прибыль и рентабельность
    net_profit = total_revenue - total_expenses
    profitability = (net_profit / total_revenue) * 100 if total_revenue else 0

    return {
        "revenue": total_revenue,
        "cost": total_cost,
        "commission": total_commission,
        "logistics": logistics_share,
        "storage": storage_share,
        "expenses": total_expenses,
        "net_profit": net_profit,
        "profitability": profitability,
        "quantity": total_quantity,
        "cost_per_item": cost_per_item,
    }
def get_comm(comission, category):
    for cat in comission["report"]:
        if cat["parentName"] == category:
            return cat["paidStorageKgvp"]

async def show_profitability_report(
    callback: types.CallbackQuery, article, state: FSMContext
):
    """Показать отчет по доходности для артикула"""
    async with state.proxy() as data:
        shop_id = data["shop"]["id"]
        shop_name = data["shop"]["name"] or f"Магазин {shop_id}"
        api_token = data["shop"]["api_token"]

    # Показываем сообщение о загрузке
    await callback.message.edit_text(
        f"📊 <b>Расчет доходности для артикула {article}</b>\n\n"
        f"Магазин: {shop_name}\n"
        "Период: последний месяц\n\n"
        "Подождите, идет расчет..."
    )

    # Рассчитываем показатели
    metrics = await calculate_profitability_for_article(article, shop_id, api_token)

    if not metrics:
        await callback.message.edit_text(
            f"❌ <b>Не удалось рассчитать доходность для артикула {article}</b>\n\n"
            "Возможные причины:\n"
            "1. Нет данных о продажах за последний месяц\n"
            "2. Не загружена себестоимость товара\n"
            "3. Проблемы с подключением к WB API"
        )
        return

    # Определяем уровень доходности
    profitability = metrics["profitability"]
    level = get_profitability_level(profitability)

    # Форматируем отчет
    text = (
        f"📊 <b>Оценка доходности: {level['name']}</b>\n\n"
        f"<b>Артикул:</b> {article}\n"
        f"<b>Магазин:</b> {shop_name}\n"
        f"<b>Период:</b> последний месяц\n\n"
        "<u>Финансовые показатели:</u>\n"
        f"💰 Выручка: {metrics['revenue']:.2f} руб.\n"
        f"📦 Продано: {metrics['quantity']} шт.\n"
        f"🏷️ Себестоимость: {metrics['cost_per_item']:.2f} руб./шт. (Итого: {metrics['cost']:.2f} руб.)\n"
        f"📊 Комиссии WB: {metrics['commission']:.2f} руб.\n"
        f"🚚 Логистика: {metrics['logistics']:.2f} руб.\n"
        f"🏭 Хранение: {metrics['storage']:.2f} руб.\n"
        f"💵 Чистая прибыль: {metrics['net_profit']:.2f} руб.\n"
        f"📈 Рентабельность: <b>{profitability:.1f}%</b>\n\n"
        f"<u>Характеристика:</u>\n{level['characteristics']}\n\n"
        f"<u>Основные причины:</u>\n{level['reasons']}\n\n"
        f"<u>Вывод:</u>\n{level['conclusion']}\n\n"
        "<u>Рекомендации:</u>\n"
    )

    # Добавляем рекомендации
    for i, recommendation in enumerate(level["recommendations"]):
        text += f"{i+1}. {recommendation}\n"

    text += f"\n<u>Действие:</u>\n🚀 <b>{level['action']}</b>"

    keyboard = InlineKeyboardMarkup()
    keyboard.add(
        InlineKeyboardButton(
            "🔙 К выбору артикула", callback_data="profitability_estimation"
        )
    )
    keyboard.add(InlineKeyboardButton("📊 В меню аналитики", callback_data="analytics"))

    await callback.message.edit_text(text, reply_markup=keyboard)


async def get_top_profitable_products(api_token: str, shop_id: int):
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=30)
    report = await fetch_full_report(api_token, start_date, end_date)
    if not report:
        return []

    session = sessionmaker()(bind=engine)
    try:
        products = {}

        for item in report:
            article = item.get("sa_name")
            if not article:
                continue

            quantity = item.get("quantity", 0)
            revenue = item.get("retail_price_withdisc_rub", 0) * quantity
            commission = (
                item.get("ppvz_sales_commission", 0)
                + item.get("ppvz_vw", 0)
                + item.get("ppvz_vw_nds", 0)
            )
            logistics = item.get("delivery_rub", 0) / len(report) * quantity + item.get("rebill_logistic_cost", 0) / len(report) * quantity
            storage = item.get("storage_fee", 0) / len(report) * quantity

            product_cost = (
                session.query(ProductCost)
                .filter(ProductCost.shop_id == shop_id, ProductCost.article == article)
                .first()
            )
            cost = product_cost.cost if product_cost else 0

            profit = revenue - (cost * quantity + commission + logistics + storage)

            if article not in products:
                products[article] = {
                    "revenue": 0,
                    "cost": cost,
                    "quantity": 0,
                    "profit": 0,
                }

            products[article]["revenue"] += revenue
            products[article]["quantity"] += quantity
            products[article]["profit"] += profit

        sorted_products = sorted(
            products.items(), key=lambda x: x[1]["profit"], reverse=True
        )

        return sorted_products[:5]

    except Exception as e:
        logger.error(f"Ошибка расчета топ-5 товаров: {e}")
        return []
    finally:
        session.close()


async def top5_products_callback(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        if "shop" not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return

        shop_id = data["shop"]["id"]
        shop_name = data["shop"]["name"] or f"Магазин {shop_id}"
        api_token = data["shop"]["api_token"]

    await callback.message.edit_text(
        "⏳ <b>Расчет топ-5 самых прибыльных товаров</b>\n\n"
        f"Магазин: {shop_name}\n"
        "Период: последний месяц\n\n"
        "Подождите, идет расчет..."
    )

    top_products = await get_top_profitable_products(api_token, shop_id)

    if not top_products:
        await callback.message.edit_text(
            "❌ <b>Не удалось получить данные</b>\n\n"
            "Проверьте, что:\n"
            "1. У вас есть подключение к интернету\n"
            "2. API-токен WB действителен\n"
            "3. Загружены данные себестоимости"
        )
        return

    text = (
        f"🏆 <b>Топ-5 самых прибыльных товаров</b>\n\n"
        f"Магазин: {shop_name}\n"
        "Период: последний месяц\n\n"
    )

    for i, (article, data) in enumerate(top_products):
        profit = data["profit"]
        revenue = data["revenue"]
        quantity = data["quantity"]
        cost = data["cost"]

        text += (
            f"{i+1}. <b>{article}</b>\n"
            f"   Прибыль: {profit:.2f} руб.\n"
            f"   Выручка: {revenue:.2f} руб.\n"
            f"   Продано: {quantity} шт.\n"
            f"   Себестоимость: {cost:.2f} руб./шт.\n\n"
        )

    text += "<i>Примечание: расчет включает себестоимость, комиссии, логистику и хранение</i>"

    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("🔙 Назад", callback_data="analytics"))

    await callback.message.edit_text(text, reply_markup=keyboard)


async def what_if_simulator_callback(callback: types.CallbackQuery, state: FSMContext):
    # Проверяем выбран ли магазин
    async with state.proxy() as data:
        if "shop" not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return
        shop_id = data["shop"]["id"]

    # Сохраняем контекст для пагинации
    async with state.proxy() as data:
        data["analytics_type"] = "what_if"
        data["article_page"] = 0

    await show_articles_page(callback, state)


async def show_articles_page(callback: types.CallbackQuery, state: FSMContext):
    session = sessionmaker()(bind=engine)
    try:
        async with state.proxy() as data:
            shop_id = data["shop"]["id"]
            page = data["article_page"]
            analytics_type = data["analytics_type"]

        articles = (
            session.query(Order.nmId).filter(Order.shop_id == shop_id).distinct().all()
        )

        articles = [art[0] for art in articles]

        if not articles:
            await callback.answer("❌ Нет данных по артикулам", show_alert=True)
            return
        items_per_page = 7
        total_pages = math.ceil(len(articles) / items_per_page)
        start_idx = page * items_per_page
        page_articles = articles[start_idx : start_idx + items_per_page]

        title = (
            "📊 Оценка доходности"
            if analytics_type == "profitability"
            else "🔮 Симулятор «А что если?»"
        )
        text = f"{title}\n\nВыберите артикул (страница {page + 1}/{total_pages}):"

        keyboard = InlineKeyboardMarkup(row_width=1)

        for article in page_articles:
            keyboard.add(
                InlineKeyboardButton(article, callback_data=f"select_article_{article}")
            )

        pagination_row = []
        if page > 0:
            pagination_row.append(
                InlineKeyboardButton("⬅️ Назад", callback_data="prev_articles_page")
            )
        if start_idx + items_per_page < len(articles):
            pagination_row.append(
                InlineKeyboardButton("Вперед ➡️", callback_data="next_articles_page")
            )

        if pagination_row:
            keyboard.row(*pagination_row)

        keyboard.add(InlineKeyboardButton("🔙 Назад", callback_data="back_to_analytics"))
        await callback.message.delete()
        await callback.message.answer(text, reply_markup=keyboard)
        await AnalyticsStates.waiting_for_article.set()
    finally:
        session.close()


async def handle_articles_pagination(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        page = data["article_page"]
        if callback.data == "prev_articles_page":
            data["article_page"] = max(0, page - 1)
        else:
            data["article_page"] = page + 1

    await show_articles_page(callback, state)


async def select_article_callback(callback: types.CallbackQuery, state: FSMContext):
    article = callback.data.split("_", 2)[2]

    async with state.proxy() as data:
        analytics_type = data["analytics_type"]
        shop_id = data["shop"]["id"]
        data["selected_article"] = article

    if analytics_type == "profitability":
        await show_profitability_report(callback, int(article), state)
    else:
        await callback.message.edit_text(
            "🔮 <b>Симулятор «А что если?»</b>\n\n"
            f"Выбран артикул: <b>{article}</b>\n\n"
            "Введите новую цену и новую себестоимость через запятую.\n"
            "Формат: <code>цена, себестоимость</code>\n"
            "Например: <code>1200, 800</code>"
        )
        await AnalyticsStates.waiting_for_price_and_cost.set()


async def what_if_simulator_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик симулятора сценариев"""
    # Проверяем выбран ли магазин
    async with state.proxy() as data:
        if "shop" not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return

    # Сохраняем контекст для пагинации
    async with state.proxy() as data:
        data["analytics_type"] = "what_if"
        data["article_page"] = 0

    await show_articles_page(callback, state)


async def process_price_and_cost(message: types.Message, state: FSMContext):
    """Обработка ввода цены и себестоимости"""
    try:
        # Пытаемся разобрать ввод
        input_text = message.text.strip()

        # Проверяем два формата: через запятую и через пробел
        if "," in input_text:
            parts = input_text.split(",")
        else:
            parts = input_text.split()

        if len(parts) != 2:
            raise ValueError

        new_price = float(parts[0].strip())
        new_cost = float(parts[1].strip())

        async with state.proxy() as data:
            article = data["selected_article"]
            shop_id = data["shop"]["id"]
            shop_name = data["shop"]["name"] or f"Магазин {shop_id}"
            api_token = data["shop"]["api_token"]

        # Получаем исторические данные
        end_date = datetime.utcnow()
        start_date = end_date - timedelta(days=30)
        report = await fetch_full_report(api_token, start_date, end_date)

        if not report:
            await message.answer(
                "❌ Не удалось получить исторические данные для расчета"
            )
            return

        # Фильтруем данные по выбранному артикулу
        article_data = [item for item in report if item.get("sa_name") == article]

        if not article_data:
            await message.answer(
                f"❌ Нет данных по артикулу {article} за последний месяц"
            )
            return

        # Рассчитываем текущие показатели
        current_quantity = sum(item.get("quantity", 0) for item in article_data)
        current_revenue = sum(
            item.get("retail_price_withdisc_rub", 0) for item in article_data
        )
        current_commission = sum(
            item.get("ppvz_sales_commission", 0)
            + item.get("ppvz_vw", 0)
            + item.get("ppvz_vw_nds", 0)
            for item in article_data
        )

        # Получаем текущую себестоимость
        session = sessionmaker()(bind=engine)
        try:
            product_cost = (
                session.query(ProductCost)
                .filter(
                    ProductCost.shop_id == shop_id, ProductCost.article == str(article)
                )
                .first()
            )
            current_cost = product_cost.cost if product_cost else 0
        finally:
            session.close()

        current_profit = (
            current_revenue - current_commission - (current_cost * current_quantity)
        )

        # Рассчитываем прогноз
        forecast_revenue = new_price * current_quantity
        forecast_profit = (
            forecast_revenue - current_commission - (new_cost * current_quantity)
        )

        # Формируем результат
        text = (
            f"🔮 <b>Симулятор «А что если?» для артикула {article}</b>\n\n"
            f"<b>Исторические данные (за последний месяц):</b>\n"
            f"📦 Продано: {current_quantity} шт.\n"
            f"💰 Выручка: {current_revenue:.2f} руб.\n"
            f"💵 Прибыль: {current_profit:.2f} руб.\n"
            f"🏷️ Текущая цена: {current_revenue / current_quantity:.2f} руб./шт.\n"
            f"📊 Текущая себестоимость: {current_cost:.2f} руб./шт.\n\n"
            f"<b>Прогноз при новых параметрах:</b>\n"
            f"🆕 Новая цена: {new_price:.2f} руб./шт.\n"
            f"🆕 Новая себестоимость: {new_cost:.2f} руб./шт.\n"
            f"📈 Прогнозируемая выручка: {forecast_revenue:.2f} руб.\n"
            f"📊 Прогнозируемая прибыль: {forecast_profit:.2f} руб.\n\n"
            f"<b>Изменение:</b>\n"
            f"💰 Выручка: {forecast_revenue - current_revenue:+.2f} руб. "
            f"({(forecast_revenue / current_revenue - 1) * 100 if current_revenue else 0:+.1f}%)\n"
            f"💵 Прибыль: {forecast_profit - current_profit:+.2f} руб. "
            f"({(forecast_profit / current_profit - 1) * 100 if current_profit else 0:+.1f}%)\n\n"
            "<i>Примечание: прогноз основан на историческом количестве продаж без учета изменения спроса</i>"
        )

        keyboard = InlineKeyboardMarkup()
        keyboard.add(
            InlineKeyboardButton("🔄 Новый расчет", callback_data="what_if_simulator")
        )
        keyboard.add(
            InlineKeyboardButton("🔙 В меню аналитики", callback_data="analytics")
        )

        await message.answer(text, reply_markup=keyboard)
        await state.finish()

    except (ValueError, IndexError):
        await message.answer(
            "❌ Неверный формат. Пожалуйста, введите цену и себестоимость через запятую или пробел.\n"
            "Пример: <code>1200, 800</code> или <code>1200 800</code>"
        )
    except Exception as e:
        logger.error(f"Ошибка в симуляторе: {e}")
        await message.answer(
            "❌ Произошла ошибка при расчете прогноза. Попробуйте позже."
        )
        await state.finish()


async def product_analytics_callback(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        if "shop" not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return

        shop_id = data["shop"]["id"]
        shop_name = data["shop"]["name"] or f"Магазин {shop_id}"
        api_token = data["shop"]["api_token"]

    # Показываем сообщение о загрузкеc
    await callback.message.delete()
    message2 = await callback.message.answer(
        "<b>Генерация товарной аналитики</b>\n\n"
        f"Магазин: {shop_name}\n"
        "Период: последний месяц\n\n"
        "Подождите, идет сбор и обработка данных..."
    )

    # Получаем данные и генерируем отчет
    try:
        wb = await generate_product_analytics_report(api_token, shop_id)
        if not wb:
            await message2.edit_text(
                "❌ <b>Не удалось сгенерировать отчет</b>\n\n"
                "Возможные причины:\n"
                "1. Нет данных о продажах за последний месяц\n"
                "2. Проблемы с подключением к WB API\n"
                "3. Отсутствуют данные себестоимости"
            )
            return

        # Сохраняем в буфер
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Отправляем файл
        await message2.delete()
        file = InputFile(file_stream, filename=f"товарная_аналитика_{shop_name}.xlsx")
        await callback.message.answer_document(
            file,
            caption=f"Товарная аналитика за последний месяц\nМагазин: {shop_name}",
        )

    except Exception as e:
        logger.error(f"Ошибка генерации отчета: {e}")
        await callback.message.edit_text(
            "❌ Произошла ошибка при генерации отчета. Попробуйте позже."
        )


async def generate_product_analytics_report(api_token: str, shop_id: int):
    """Генерация Excel-отчета с товарной аналитикой"""
    end_date = datetime.utcnow()
    start_date = datetime(year=end_date.year, month=end_date.month, day=1)
    star = datetime.today() - timedelta(days=datetime.today().isoweekday())
    week_start = datetime(star.year, star.month, star.day + 1, 0, 0)

    # Получаем отчет за последний месяц
    session = sessionmaker(bind=engine)()
    report = (
        session.query(CashedShopData)
        .filter(CashedShopData.shop_id == shop_id)
        .first()
        .cashed_month
    )
    session.close()
    if not report:
        return None

    # Создаем Excel-книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товарная аналитика"

    # Заголовки столбцов
    headers = [
        "Наименование",
        "Артикул",
        "Заказы (шт)",
        "Продажи (шт)",
        "Возвраты (шт)",
        "Отмены (шт)",
        "Продажи (руб)",
        "Возвраты (руб)",
        "Выручка",
        "Итого продаж (шт)",
        "% выкупа",
        "Комиссия (руб)",
        "% комиссии",
        "Логистика (руб)",
        # "Обратная логистика (руб)",
        "Логистика на ед",
        "% логистики",
        "Все удержания",
        "% удержаний",
        "Налог",
        "Прибыль без рекламы",
        "Реклама",
        "Удержания",
        "Чистая прибыль с рекламой",
        "Рентабельность CPM",
    ]

    # Добавляем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
        )
    regular_expenses = 0
    days_in_period = (end_date - start_date).days + 1
    for expense in session.query(RegularExpense).filter(
            RegularExpense.shop_id == shop_id
    ):
        if expense.frequency == RegularExpenseFrequency.DAILY:
            regular_expenses += expense.amount * days_in_period
        elif expense.frequency == RegularExpenseFrequency.WEEKLY:
            regular_expenses += expense.amount * (days_in_period / 7)
        elif expense.frequency == RegularExpenseFrequency.MONTHLY:
            regular_expenses += expense.amount * (days_in_period / 30)

    # Собираем данные по артикулам
    articles_data = {}
    for item in report:
        article = item.get("sa_name")
        # if not article:
        #     continue
        if not article:

            if item.get("nm_id", 0):
                print(item)
                print(articles_data)
                print("\n\n")
                for article2, item2 in articles_data.items():
                    if item2.get("nm_id") == item.get("nm_id"):
                        article = article2
        if not article:
            continue

        if article not in articles_data:
            articles_data[article] = {
                "subject_name": item.get("subject_name", ""),
                "orders": 0,
                "sales": 0,
                "returns": 0,
                "cancellations": 0,
                "sales_rub": 0,
                "returns_rub": 0,
                "commission": 0,
                "logistics": 0,
                "storage": 0,
                "return_logistics": 0,  # Обратная логистика (пока нет данных)
                "nm_id": item.get("nm_id", 0),
                "deduction": 0
            }

        doc_type = item.get("doc_type_name", "")
        quantity = item.get("quantity", 0)
        price = item.get("retail_price_withdisc_rub", 0)
        retail_price = item.get("retail_price_withdisc_rub", 0)

        if "продажа" in doc_type.lower() or "sale" in doc_type.lower():
            articles_data[article]["sales"] += quantity
            articles_data[article]["sales_rub"] += price*quantity
        elif "возврат" in doc_type.lower() or "return" in doc_type.lower():
            articles_data[article]["returns"] += quantity
            articles_data[article]["returns_rub"] += price
        elif "отмена" in doc_type.lower() or "cancellation" in doc_type.lower():
            articles_data[article]["cancellations"] += quantity
        articles_data[article]["deduction"] += item['deduction']

        articles_data[article]["commission"] += retail_price - item.get("ppvz_for_pay", 0)
        articles_data[article]["commission"] -= item.get("ppvz_reward", 0)
        articles_data[article]["commission"] -= item.get("ppvz_sales_commission", 0)

        articles_data[article]["logistics"] += item.get("delivery_rub", 0)
        articles_data[article]["storage"] += item["storage_fee"]
        # print(item["storage_fee"])

        articles_data[article]["orders"] += quantity

    print(week_start)
    orders = (
        session.query(Order)
        .filter(Order.is_bouhght.is_(True))
        .filter(Order.date >= week_start)
        .filter(Order.isCancel.is_(False))
        .filter(Order.shop_id == shop_id)
        .all()
    )
    # for article, item in articles_data.items():
    #     print(item.get("commission"))

    for order in orders:
        if order.supplierArticle not in articles_data:
            articles_data[order.supplierArticle] = {
                "subject_name": order.supplierArticle,
                "orders": 0,
                "sales": 0,
                "returns": 0,
                "cancellations": 0,
                "sales_rub": 0,
                "returns_rub": 0,
                "commission": 0,
                "logistics": 0,
                "storage": 0,
                "return_logistics": 0,
                "nm_id": order.nmId,
                "deduction": 0
            }
        articles_data[order.supplierArticle]["sales_rub"] += order.priceWithDisc
        articles_data[order.supplierArticle]["sales"] += 1
        articles_data[order.supplierArticle]["orders"] += 1
        articles_data[order.supplierArticle]["commission"] += order.priceWithDisc - order.forPay
    amount_articles = len(articles_data)

    for item in report:
        if item.get("nm_id", 0) == 0:
            if item.get('bonus_type_name', '') == "Оказание услуг «ВБ.Продвижение»":
                continue
            if item.get("ppvz_reward", 0):
                for item2 in report:
                    if item2.get("srid") == item.get("srid"):
                        if item2.get("sa_name"):
                            articles_data[item2.get("sa_name")]["commission"] -= item.get("ppvz_reward")
                            break
            deduction = item.get("deduction", 0)/amount_articles
            storage = item.get("storage_fee", 0)/amount_articles
            for article, data in articles_data.items():
                data["deduction"] += deduction
                data["storage"] += storage

    regular_expenses_for_article = regular_expenses/amount_articles
    session = sessionmaker()(bind=engine)
    try:
        product_costs = (
            session.query(ProductCost).filter(ProductCost.shop_id == shop_id).all()
        )
        # print(product_costs)
        cost_map = {pc.article: pc.cost for pc in product_costs}
    finally:
        session.close()
    # print(cost_map, "\n\n|^ COST MAP")
    # Налоговая ставка
    session = sessionmaker()(bind=engine)
    try:
        tax_setting = (
            session.query(TaxSystemSetting)
            .filter(TaxSystemSetting.shop_id == shop_id)
            .first()
        )
        tax_rate = (
            0.06
            if tax_setting and tax_setting.tax_system == TaxSystemType.USN_6
            else 0.0
        )
    finally:
        session.close()

    # Заполняем данные в таблицу
    row_num = 2
    for article, data in articles_data.items():
        # Основные показатели
        revenue = data["sales_rub"] - data["returns_rub"]
        total_sales = data["sales"] - data["returns"]
        buyout_rate = (total_sales / data["orders"]) if data["orders"] else 0

        # Комиссии
        commission_percent = (data["commission"] / revenue) if revenue else 0

        # Логистика
        logistics_per_unit = data["logistics"] / total_sales if total_sales else 0
        logistics_percent = (data["logistics"] / revenue) if revenue else 0

        # Хранение
        storage_percent = (data["storage"] / revenue) if revenue else 0

        # Удержания
        total_deductions = (
            data["commission"]
            + data["logistics"]
            + data["return_logistics"]
            + data["storage"]
            + data["deduction"]
        )
        # print(article, data["commission"], data["logistics"], data["return_logistics"], data["storage"], data["deduction"])
        deductions_percent = (total_deductions / revenue) if revenue else 0

        # Налог
        tax = revenue * tax_rate

        # Себестоимость
        cost_per_item = cost_map.get(article, 0)
        total_cost = cost_per_item * total_sales

        # Прибыль
        profit_without_ads = (
            revenue - abs(total_cost) - abs(total_deductions) - abs(tax) - abs(regular_expenses_for_article)
        )
        profit_with_ads = profit_without_ads  # Рекламные расходы не учитываем
        adverisement = sum(
            i.amount
            for i in session.query(Advertisement)
            .filter(Advertisement.nmId == int(data["nm_id"]))
            .filter(Advertisement.date >= start_date)
            .all()
        )
        # print(adverisement, int(data["nm_id"]), start_date, len(session.query(Advertisement).filter(Advertisement.nmId == int(data["nm_id"])).filter(Advertisement.date >= start_date).all()))
        # data2r3 = session.query(Advertisement).filter(Advertisement.nmId == int(data["nm_id"])).filter(Advertisement.date >= start_date)
        # print(data2r3)
        penalty = sum(
            i.sum
            for i in session.query(Penalty)
            .filter(Penalty.nm_id == data["nm_id"])
            .filter(Penalty.date >= datetime.now() - timedelta(days=30))
            .all()
        )
        # print((data["nm_id"]), adverisement, penalty)
        profit_with_ads = (
            revenue
            - abs(total_cost)
            - abs(total_deductions)
            - abs(tax)
            - abs(adverisement)
            - abs(penalty)
            - abs(regular_expenses_for_article)
        )
        # Рентабельность
        profitability_cpm = (profit_without_ads / total_cost) * 100 if total_cost else 0

        # Заполняем строку
        ws.cell(row=row_num, column=1, value=data["subject_name"])
        ws.cell(row=row_num, column=2, value=article)
        ws.cell(row=row_num, column=3, value=abs(data["orders"]))
        ws.cell(row=row_num, column=4, value=abs(data["sales"]))
        ws.cell(row=row_num, column=5, value=abs(data["returns"]))
        ws.cell(row=row_num, column=6, value=abs(data["cancellations"]))
        ws.cell(row=row_num, column=7, value=abs(data["sales_rub"]))
        ws.cell(row=row_num, column=8, value=abs(data["returns_rub"]))
        ws.cell(row=row_num, column=9, value=abs(revenue))
        ws.cell(row=row_num, column=10, value=abs(total_sales))
        ws.cell(row=row_num, column=11, value=abs(buyout_rate))
        ws.cell(row=row_num, column=12, value=data["commission"])
        ws.cell(row=row_num, column=13, value=abs(commission_percent))
        ws.cell(row=row_num, column=14, value=abs(data["logistics"]+data["return_logistics"]))
        # ws.cell(row=row_num, column=15, value=abs())
        ws.cell(row=row_num, column=15, value=abs(logistics_per_unit))
        ws.cell(row=row_num, column=16, value=abs(logistics_percent))
        ws.cell(row=row_num, column=17, value=abs(total_deductions))
        ws.cell(row=row_num, column=18, value=abs(deductions_percent))
        ws.cell(row=row_num, column=19, value=abs(tax))
        ws.cell(row=row_num, column=20, value=int(profit_without_ads))
        ws.cell(row=row_num, column=21, value=adverisement)
        ws.cell(row=row_num, column=22, value=data["deduction"])
        ws.cell(row=row_num, column=23, value=profit_with_ads)
        ws.cell(row=row_num, column=24, value=profitability_cpm)

        row_num += 1

    # Форматирование
    apply_excel_formatting(ws)

    return wb


def apply_excel_formatting(ws):
    """Применяет форматирование к Excel-листу"""
    # Устанавливаем ширину столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Форматирование чисел
    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=3, max_col=ws.max_column
    ):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                # Проценты
                if cell.column_letter in ["K", "M", "P", "R"]:
                    cell.number_format = "0.00%"
                elif (
                    cell.column >= 7
                    and cell.column <= 24
                    and cell.column not in [10, 11]
                ):
                    cell.number_format = "#,##0.00"
                else:
                    cell.number_format = "#,##0"

    # Границы
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border

    # Выравнивание заголовков
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Фиксируем заголовки
    ws.freeze_panes = "A2"


async def back_to_analytics(callback: types.CallbackQuery, state: FSMContext):
    await analytics_callback(callback)

async def finances_handler(callback: types.CallbackQuery, state: FSMContext):
    text = "<b>Это раздел финансов</b>\n\nЗдесь Вы можете узнать свои главные показатели по своему бизнесу.\n\n▫️ Чистая прибыль\n▫️ Сроки окупаемости с учетом всех Ваших первоначальных затрат\n▫️ Рентабельность инвестиций покажет, насколько выгоден Ваш проект и как быстро он окупается\n▫️ Годовая доходность Вашего бизнеса покажет, насколько выгоден Ваш бизнес"
    session = sessionmaker(bind=engine)()
    print(callback.from_user.id)
    user = session.query(User).filter(User.telegram_id == callback.from_user.id).first()
    session.close()

    if user:
        if user.subscription_end <= datetime.now():
            text += '\n\n⚠️ У вас закончилась подписка <b>JustProfit Premium</b>. Продлите её и Вам сразу же будут доступны все функции бота. \n\nПродлить подписку: "Главное меню" -> "Поддержка" -> "Подписка"'
            kb = InlineKeyboardMarkup()
            kb.add(InlineKeyboardButton("Меню", callback_data="main_menu"))
            await callback.message.delete()
            await callback.message.answer(text, reply_markup=kb)
            return
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("Чистая прибыль", callback_data="an_1"))
    #kb.add(InlineKeyboardButton("ROS(Рентабльность продаж)", callback_data="an_2"))
    kb.add(InlineKeyboardButton("Срок окупаемости", callback_data="an_3"))
    kb.add(InlineKeyboardButton("ROI(Рентабельность вложений)", callback_data="an_4"))
    kb.add(InlineKeyboardButton("Годовая доходность", callback_data="an_5"))
    kb.add(InlineKeyboardButton("Меню", callback_data="main_menu"))
    await callback.message.delete()
    await callback.message.answer(text, reply_markup=kb)


async def pnl_callback(callback: types.CallbackQuery, state: FSMContext):
    # Проверяем выбран ли магазин
    async with state.proxy() as data:
        if "shop" not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return

    await callback.message.edit_text(
        "📊 <b>Расчёт прибыли и убытков (P&L)</b>\n\n" "Выберите период для расчета:",
    )


# Расчет показателей на основе отчета
async def calculate_metrics_from_report(report_data, shop_id, start_date, end_date, type_data="week", calculate_current_week=True):
    session = sessionmaker()(bind=engine)
    try:
        # Основные показатели
        star = datetime.today() - timedelta(days=datetime.today().isoweekday())
        week_start = datetime(star.year, star.month, star.day + 1, 0, 0)
        print(start_date, end_date)
        revenue = 0
        logistics = 0
        storage_fee = 0
        commission = 0
        cost_of_goods = 0
        deduction = 0
        ppvz_reward = 0
        # Собираем артикулы для расчета себестоимости
        articles = {}
        for item in report_data:
            # Выручка
            if item.get('bonus_type_name', '') == "Оказание услуг «ВБ.Продвижение»":
                continue
            retail_price = item.get("retail_price_withdisc_rub", 0)
            quantity = item.get("quantity", 0)
            revenue += retail_price * quantity

            # Логистика
            logistics += item.get("delivery_rub", 0)
            # logistics += item.get("rebill_logistic_cost", 0)

            # Хранение
            storage_fee += item.get("storage_fee", 0)

            # Комиссия WB
            # print(item.get('ppvz_sales_commission'), item.get('ppvz_vw'), item.get('ppvz_vw_nds'))
            commission += retail_price - item.get("ppvz_for_pay")
            commission -= item.get("ppvz_reward", 0)
            ppvz_reward += item.get("ppvz_reward", 0)
            commission -= item.get("ppvz_sales_commission", 0)
            deduction += item.get("deduction", 0)

            # Собираем данные для себестоимости
            article = item.get("nm_id")
            if article:
                if article not in articles:
                    articles[article] = 0
                articles[article] += quantity
        # print(shop_id)
        print(commission, ppvz_reward)
        if calculate_current_week:
            orders = (
                session.query(Order)
                .filter(Order.is_bouhght.is_(True))
                .filter(Order.date >= week_start)
                .filter(Order.isCancel.is_(False))
                .filter(Order.shop_id == shop_id)
                .all()
            )
            for order in orders:
                if not order.isCancel:
                    revenue += order.priceWithDisc
                    commission += order.priceWithDisc - order.forPay
                    article = order.nmId
                    if article not in articles:
                        articles[article] = 0
                    articles[article] += 1


        # Себестоимость
        for article, quantity in articles.items():
            print(article, quantity)
            try:
                supp_article = session.query(Order).filter(Order.nmId == int(article)).first().supplierArticle
                product_cost = (
                    session.query(ProductCost)
                    .filter(ProductCost.shop_id == shop_id, ProductCost.article == supp_article)
                    .first()
                )
                if product_cost:
                    cost_of_goods += product_cost.cost * quantity
            except:
                pass
        # Налоговая ставка
        tax_setting = (
            session.query(TaxSystemSetting)
            .filter(TaxSystemSetting.shop_id == shop_id)
            .first()
        )

        # TAX RATE TAX RATE TAX RATE

        tax_rate = (
            0.06
            if tax_setting and tax_setting.tax_system == TaxSystemType.USN_6
            else 0.0
        )
        tax = revenue * tax_rate

        # Регулярные затраты за период
        regular_expenses = 0
        days_in_period = (end_date - start_date).days + 1
        # print(start_date, end_date, days_in_period)
        for expense in session.query(RegularExpense).filter(RegularExpense.shop_id == shop_id):
            if type_data == "week":
                if expense.frequency == RegularExpenseFrequency.WEEKLY:
                    regular_expenses += expense.amount * 1
                if expense.frequency == RegularExpenseFrequency.DAILY:
                    regular_expenses += expense.amount * 7
            if type_data == "month":
                if expense.frequency == RegularExpenseFrequency.DAILY:
                    regular_expenses += expense.amount * 30
                if expense.frequency == RegularExpenseFrequency.WEEKLY:
                    regular_expenses += expense.amount * 4
                if expense.frequency == RegularExpenseFrequency.MONTHLY:
                    regular_expenses += expense.amount * 1
            if type_data == "year":
                if expense.frequency == RegularExpenseFrequency.DAILY:
                    regular_expenses += expense.amount * 365
                if expense.frequency == RegularExpenseFrequency.WEEKLY:
                    regular_expenses += expense.amount * 52
                if expense.frequency == RegularExpenseFrequency.MONTHLY:
                    regular_expenses += expense.amount * 12


        advert = sum(
            i.amount
            for i in session.query(Advertisement)
            .filter(Advertisement.shop_id == shop_id)
            .filter(Advertisement.date >= start_date)
            .all()
        )
        # print(advert)

        # Удержания
        stops = sum(
            i.sum
            for i in session.query(Penalty).filter(Penalty.date >= start_date).filter(Penalty.shop_id == shop_id).all()
        )

        # Чистая прибыль
        net_profit = revenue - (
            commission
            + logistics
            + storage_fee
            + tax
            + cost_of_goods
            + regular_expenses
            + stops
            + deduction
            + advert
        )

        # Рентабельность
        profitability = (net_profit / revenue) * 100 if revenue > 0 else 0

        # Рекламные затраты

        # Разовые затраты (инвестиционные)
        one_time_expenses = (
            session.query(OneTimeExpense)
            .filter(OneTimeExpense.shop_id == shop_id)
            .all()
        )
        total_one_time = sum(expense.amount for expense in one_time_expenses)

        # НОВЫЙ РАСЧЕТ СРОКА ОКУПАЕМОСТИ С ЛИНЕЙНОЙ РЕГРЕССИЕЙ
        import numpy as np
        from collections import defaultdict

        # 1. Собираем прибыль по каждому месяцу
        profits_by_month = defaultdict(float)

        orders = (
            session.query(Order)
            .filter(Order.shop_id == shop_id)
            .filter(Order.is_bouhght.is_(True))
            .filter(Order.isCancel.is_(False))
            .order_by(Order.date.asc())
            .all()
        )

        for order in orders:
            month_key = order.date.strftime("%Y-%m")
            profit = order.priceWithDisc - (order.forPay or 0)
            profits_by_month[month_key] += profit

        if profits_by_month and total_one_time > 0:
            # 2. Преобразуем данные в формат для регрессии
            sorted_months = sorted(profits_by_month.items())
            X = np.arange(len(sorted_months)).reshape(-1, 1)  # Месяцы: 0, 1, 2, ...
            y = np.array([profit for _, profit in sorted_months])  # Прибыль

            # 3. Строим линейную регрессию y = kx + b
            coeffs = np.polyfit(X.flatten(), y, 1)  # 1 степень = линейная
            k, b = coeffs

            # 4. Прогнозируем прибыль на следующие месяцы, пока не окупится
            total_cumulative_profit = sum(y)
            month_idx = len(X)  # начинаем прогноз после последнего известного месяца
            while total_cumulative_profit < total_one_time and month_idx < 120:  # ограничение 10 лет
                predicted_profit = k * month_idx + b
                predicted_profit = max(predicted_profit, 0)  # чтобы не было отрицательной прибыли
                total_cumulative_profit += predicted_profit
                month_idx += 1

            if total_cumulative_profit >= total_one_time:
                payback_period = f"{month_idx} месяцев"
            else:
                payback_period = "не определен (даже при прогнозе на 10 лет)"
        else:
            # Старый расчет как fallback
            payback_period = "не определен, добавьте, пожалуйста, внешние расходы"
            if net_profit > 0 and total_one_time > 0:
                months = total_one_time / net_profit
                payback_period = f"{months:.1f} месяцев"

        # ROI
        roi = "не определен"
        if total_one_time > 0:
            roi_value = (net_profit / total_one_time) * 100
            roi = f"{roi_value:.1f}%"
            if roi_value > 100:
                roi += " ✅ Поздравляем, вы окупили вложения!"

        try:
            ros_value = (net_profit / revenue) * 100
        except:
            ros_value = 0
        return {
            "revenue": revenue,
            "commission": commission,
            "logistics": logistics,
            "storage": storage_fee,
            "cost_of_goods": cost_of_goods,
            "tax": tax,
            "regular_expenses": regular_expenses,
            "net_profit": net_profit,
            "profitability": profitability,
            "payback_period": payback_period,
            "roi": roi,
            "total_one_time": total_one_time,
            "advert": advert,
            "stops": stops,
            "deduction": deduction,
            "ros": ros_value,
        }
    finally:
        session.close()


async def select_anal_period_callback(callback: types.CallbackQuery, state: FSMContext):

    
    period_type = callback.data.split("_")[1]  # day, week, month, year, custom
    await callback.message.delete()
    message = await callback.message.answer(
        text="Производим расчёт данных, пожалуйста, подождите\n\n"
             "‼️ Важно: <u>необходимо подождать около 10 секунд для полного завершения расчёта "
             "(иногда это может занять больше времени, но не более 2 минут)</u>"
    )

    async with state.proxy() as data:
        shop_id = data["shop"]["id"]
        shop_name = data["shop"]["name"] or f"Магазин {shop_id}"
        api_token = data["shop"]["api_token"]
        an_type = data["an_type"]

        now = datetime.now()

        # Определяем даты периода
        if period_type == "custom" or period_type.startswith("custom_"):
            # Проверяем, есть ли кастомные даты
            if data.get("custom_period") and data.get("custom_start_date") and data.get("custom_end_date"):
                current_start = data["custom_start_date"]
                current_end = data["custom_end_date"]
                type_datalol = data.get("period_size", "custom")
                period_name = f"{current_start.strftime('%d.%m')}-{current_end.strftime('%d.%m.%Y')}"
            else:
                # Фоллбек на текущий день, если кастомные даты не заданы
                current_start = datetime(now.year, now.month, now.day)
                current_end = now
                type_datalol = "day"
                period_name = f"{current_start.strftime('%d.%m')}-{now.strftime('%d.%m')}"
        elif period_type == "week":
            start_week = now - timedelta(days=now.isoweekday() - 1)
            current_start = datetime(start_week.year, start_week.month, start_week.day)
            current_end = now
            type_datalol = "week"
            period_name = f"{current_start.strftime('%d.%m')}-{now.strftime('%d.%m')}"
        elif period_type == "month":
            current_start = datetime(now.year, now.month, 1)
            current_end = now
            type_datalol = "month"
            period_name = f"{current_start.strftime('%d.%m')}-{now.strftime('%d.%m')}"
        elif period_type == "year":
            current_start = datetime(now.year, 1, 1)
            current_end = now
            type_datalol = "year"
            period_name = f"{current_start.strftime('%d.%m')}-{now.strftime('%d.%m')}"
        else:
            # По умолчанию - неделя
            start_week = now - timedelta(days=now.isoweekday() - 1)
            current_start = datetime(start_week.year, start_week.month, start_week.day)
            current_end = now
            type_datalol = "week"
            period_name = f"{current_start.strftime('%d.%m')}-{now.strftime('%d.%m')}"

    # Загружаем или получаем кэш
    session = sessionmaker(bind=engine)()
    cashed = session.query(CashedShopData).filter(CashedShopData.shop_id == shop_id).first()
    if cashed is None:
        # Ваш код загрузки данных из API и кэширования (оставляем без изменений)
        # ...
        session.close()
        session = sessionmaker(bind=engine)()
        cashed = session.query(CashedShopData).filter(CashedShopData.shop_id == shop_id).first()

    # Фильтруем отчет по периоду
    if period_type == "custom" or period_type.startswith("custom_"):
        current_report = []
        for item in cashed.cashed_all or []:
            try:
                sale_date = datetime.strptime(item.get("sale_dt", "")[:10], "%Y-%m-%d")
                if current_start <= sale_date <= current_end:
                    current_report.append(item)
            except Exception:
                continue
    else:
        if period_type == "week":
            current_report = cashed.cashed_week or []
        elif period_type == "month":
            current_report = cashed.cashed_month or []
        elif period_type == "year":
            current_report = cashed.cashed_year or []
        else:
            current_report = cashed.cashed_all or []

    # Если нет данных и период не неделя — предупреждаем
    if not current_report and period_type != "week":
        await callback.answer(
            "❌ Не удалось получить данные за текущий период, подождите около 1-2 минуты и попробуйте снова",
            show_alert=True,
        )
        return


    async with state.proxy() as data:
        shop_id = data["shop"]["id"]
        shop_name = data["shop"]["name"] or f"Магазин {shop_id}"
        api_token = data["shop"]["api_token"]
        an_type = data["an_type"]
        type_data = int(an_type.split("_")[1])  # <-- ПЕРЕНЕСИТЕ СЮДА

               
    
        
    if type_data == 3:
        current_report = cashed.cashed_month or []
        current_start = datetime(now.year, now.month, 1)
        current_end = now
        type_datalol = "month"

    # обработка type_data 5 только если НЕ кастомный период
    if type_data == 5:
        current_report = []
        start_now = datetime.now() - timedelta(days=365)
        if cashed.cashed_all:
            for i in cashed.cashed_all:
                if isinstance(i, dict) and "sale_dt" in i:
                    try:
                        if datetime.strptime(i["sale_dt"][:10], "%Y-%m-%d") >= start_now:
                            current_report.append(i)
                    except (ValueError, TypeError):
                        continue


    if not current_report and period_type != "week":
        await callback.answer(
            "❌ Не удалось получить данные за текущий период, подождите около 1-2 минуты и попробуйте снова",
            show_alert=True,
        )
        return
    # print(an_type)
    logger.info(f"Create AN report. AN-type: {an_type}")
    report = cashed.cashed_all
    await message.edit_text(text="Осталось совсем чуть-чуть ...")

    # Рассчитываем показатели
    if an_type != "an_5":
        current_metrics = await calculate_metrics_from_report(
            current_report, shop_id, current_start, current_end, type_datalol
        )

    else:
        amount_good_months = 0
        net_profit = 0
        report_data = []
        type_data = "year"
        for i in range(12):
            start_now = datetime.now() - timedelta(days=31 + i * 30)
            end_now = datetime.now() - timedelta(days=i * 30)
            new_report = []
            for report_str in report:
                if (start_now <= datetime.strptime(report_str["sale_dt"][:10], "%Y-%m-%d") <= end_now):
                    new_report.append(report_str)
            if new_report != []:
                amount_good_months += 1
                report_data += new_report
            else:
                break
        start_now = datetime.now() - timedelta(days=31 + amount_good_months * 30)
        end_now = datetime.now()
        current_metrics = await calculate_metrics_from_report(
            report_data, shop_id, start_now, end_now, type_datalol
        )

    # Рассчитываем динамику
    # revenue_change = current_metrics["revenue"] - (previous_metrics["revenue"] if previous_metrics else 0)
    # profit_change = current_metrics["net_profit"] - (previous_metrics["net_profit"] if previous_metrics else 0)

    # revenue_indicator = "🟢▲" if revenue_change >= 0 else "🔴▼"
    # profit_indicator = "🟢▲" if profit_change >= 0 else "🔴▼"
    if an_type =="an_1":
        if period_type == "week":
            date_start = current_start - timedelta(days=7)
            date_end = date_start + timedelta(days=7)
            new_report = []
            for report_str in report:
                if (date_start <= datetime.strptime(report_str["sale_dt"][:10], "%Y-%m-%d") <= date_end):
                    new_report.append(report_str)
            last_metrics = await calculate_metrics_from_report(new_report, shop_id, date_start, date_end, "week", False)
        elif period_type == "month":
            date_start = current_start - timedelta(days=30)
            date_end = date_start + timedelta(days=30)
            new_report = []
            for report_str in report:
                if (date_start <= datetime.strptime(report_str["sale_dt"][:10], "%Y-%m-%d") <= date_end):
                    new_report.append(report_str)
            last_metrics = await calculate_metrics_from_report(new_report, shop_id, date_start, date_end, "month", False)
        elif period_type == "year":
            date_start = current_start - timedelta(days=365)
            date_end = date_start + timedelta(days=365)
            new_report = []
            for report_str in report:
                if (date_start <= datetime.strptime(report_str["sale_dt"][:10], "%Y-%m-%d") <= date_end):
                    new_report.append(report_str)
            last_metrics = await calculate_metrics_from_report(new_report, shop_id, date_start, date_end, "year", False)
        elif period_type == "custom":
            # Получаем кастомные даты из state
            async with state.proxy() as data:
                date_start = data.get("custom_start_date")
                date_end = data.get("custom_end_date")
            
            if date_start and date_end:
                new_report = []
                for report_str in report:
                    if (date_start <= datetime.strptime(report_str["sale_dt"][:10], "%Y-%m-%d") <= date_end):
                        new_report.append(report_str)
                last_metrics = await calculate_metrics_from_report(new_report, shop_id, date_start, date_end, "custom", False)
            else:
                # Fallback если кастомные даты не найдены
                date_start = current_start - timedelta(days=30)
                date_end = date_start + timedelta(days=30)
                new_report = []
                for report_str in report:
                    if (date_start <= datetime.strptime(report_str["sale_dt"][:10], "%Y-%m-%d") <= date_end):
                        new_report.append(report_str)
                last_metrics = await calculate_metrics_from_report(new_report, shop_id, date_start, date_end, "month", False)
            

    # Форматируем отчет
    text = ""

    #
    #         f"💵 Чистая прибыль: {current_metrics['net_profit']:.2f} руб. \n"
    #         f"📈 Рентабельность: {current_metrics['profitability']:.1f}%\n"
    #         f"⏳ Срок окупаемости: {current_metrics['payback_period']}\n"
    #         f"📊 ROI: {current_metrics['roi']}\n\n"
    async with state.proxy() as data:
        an_type = data["an_type"]
    if an_type == "an_1":
        # Блок an_1
        destanation = f"▲ {last_metrics['net_profit']:.2f} руб." if last_metrics['net_profit'] < current_metrics['net_profit'] else f"▼ {last_metrics['net_profit']:.2f} руб."
        text = (
            f"Период: <b>({period_name})</b>\n\n"
            "<u>Основные показатели:</u>\n"
            f"▫️Выручка: {current_metrics['revenue']:.2f} руб.\n"
            f"▫️Комиссии: {current_metrics['commission']:.2f} руб. <b>{current_metrics['commission']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Логистика: {current_metrics['logistics']:.2f} руб. <b>{current_metrics['logistics']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Хранение: {current_metrics['storage']:.2f} руб. <b>{current_metrics['storage']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Себестоимость: {current_metrics['cost_of_goods']:.2f} руб. <b>{current_metrics['cost_of_goods']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Налог: {current_metrics['tax']:.2f} руб. <b>{current_metrics['tax']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Регулярные затраты: {current_metrics['regular_expenses']:.2f} руб. <b>{current_metrics['regular_expenses']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Рекламные затраты: {current_metrics['advert']} руб. <b>{current_metrics['advert']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Прочие удержания: {current_metrics['deduction']} руб. <b>{current_metrics['deduction']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Штрафы: {current_metrics['stops']} руб. <b>{current_metrics['stops']/current_metrics['revenue']*100:.1f}%</b>\n\n"
            f"〽️ Чистая прибыль: {current_metrics['net_profit']:.2f} руб. ({destanation}) <b>{current_metrics['net_profit']/current_metrics['revenue']*100:.1f}%</b>\n\n"
        )

        # Блок an_3 (окупаемость)
        text += (
            "<u>Срок окупаемости:</u>\n"
            f"▫️Разовые вложения: {current_metrics['total_one_time']:.2f} руб.\n"
            f"▫️Чистая прибыль за месяц: {current_metrics['net_profit']:.2f} руб.\n"
            f"🧮 Срок окупаемости = {current_metrics['payback_period']}\n\n"
        )

        # Блок an_4 (рентабельность)
        text += (
            "<u>Рентабельность инвестиций:</u>\n"
            f"▫️Выручка: {current_metrics['revenue']:.2f} руб.\n"
            f"▫️Чистая прибыль: {current_metrics['net_profit']:.2f} руб.\n"
            f"▫️Разовые вложения: {current_metrics['total_one_time']:.2f} руб.\n"
            f"📊 ROI: {current_metrics['roi']}\n\n"
        )

        # Блок an_5 (годовая доходность)
        amount_good_months = 0
        net_profit = 0
        report_data = []
        for i in range(12):
            start_now = datetime.now() - timedelta(days=31 + i * 30)
            end_now = datetime.now() - timedelta(days=i * 30)
            new_report = []
            for report_str in report:
                if (start_now <= datetime.strptime(report_str["sale_dt"][:10], "%Y-%m-%d") <= end_now):
                    new_report.append(report_str)
            if new_report:
                amount_good_months += 1
                report_data += new_report
            else:
                break
        start_now = datetime.now() - timedelta(days=31 + amount_good_months * 30)
        end_now = datetime.now()
        metrics_for_an_5 = await calculate_metrics_from_report(
            report_data, shop_id, start_now, end_now, "year"
        )
        text += (
            "<u>Годовая доходность:</u>\n"
            f"▫️Чистая прибыль за {amount_good_months} мес.: {metrics_for_an_5['net_profit']:.2f} руб.\n"
            f"▫️Годовая доходность: {metrics_for_an_5['roi']}\n"
        )

        text += f"\n<i>Примечание: расчеты основаны на данных WB API</i>"

        await message.edit_text(text)
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("🔙 Назад", callback_data="main_menu"))
    # await message.delete()
    await message.edit_text(text, reply_markup=keyboard)
    # await state.finish()
    # kb.add(InlineKeyboardButton("Чистая прибыль", callback_data="an_1"))
    # kb.add(InlineKeyboardButton("ROS(Рентабльность продаж)", callback_data="an_2"))
    # kb.add(InlineKeyboardButton("Срок окупаемости", callback_data="an_3"))
    # kb.add(InlineKeyboardButton("ROI(Рентабельность вложений)", callback_data="an_4"))


async def anal_callback(callback: types.CallbackQuery, state: FSMContext):
    # Проверяем выбран ли магазин

    async with state.proxy() as data:
        if "shop" not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return
        print(callback.data)
        data["an_type"] = callback.data
    type_data = int(callback.data.split("_")[1])
    # print(type_data)
    text = ""
    if type_data == 1:
        text = "Чистой прибыли"
        keyboard = period_keyboard(type_data)
    elif type_data == 2:
        text = "ROS (Рентабельности продаж)"
        keyboard = period_keyboard2(type_data)
    elif type_data == 3:
        text = "Срока окупаемости"
        return await select_anal_period_callback(callback, state)

    elif type_data == 4:
        text = "Рентабельность вложений"
        return await select_anal_period_callback(callback, state)
    else:
        return await select_anal_period_callback(callback, state)
        text = "Годовой доходности"

    # print(text)
    await callback.message.delete()
    await callback.message.answer(
        f" <b>Расчёт {text}</b>\n\n" "Выберите период для расчета:",
        reply_markup=keyboard,
    )

async def custom_period_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик для кнопки 'Выбранный период' - показывает кнопки размера периода"""
    type_data = callback.data.split("_")[-1]
    
    async with state.proxy() as data:
        data["an_type"] = f"an_{type_data}"
    
    keyboard = InlineKeyboardMarkup(row_width=1)
    keyboard.add(
        InlineKeyboardButton("📅 День", callback_data=f"custom_day_{type_data}"),      # Изменено
        InlineKeyboardButton("📅 Неделя", callback_data=f"custom_week_{type_data}"),   # Изменено
        InlineKeyboardButton("📅 Месяц", callback_data=f"custom_month_{type_data}"),   # Изменено
        InlineKeyboardButton("🔙 Назад", callback_data="main_menu")
    )
    
    try:
        await callback.message.edit_text(
            " <b>Выберите размер периода для расчета</b>\n\n"
            "Укажите, за какой период вы хотите получить данные:",
            reply_markup=keyboard
        )
    except MessageNotModified:
        # Если сообщение не изменилось, просто отвечаем на callback
        await callback.answer()
    except Exception as e:
        # Для других ошибок логируем и отвечаем на callback
        print(f"Error in custom_period_callback: {e}")
        await callback.answer()

async def custom_period_size_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик выбора размера периода - показывает календарь"""
    parts = callback.data.split("_")
    period_size = parts[1]  # day, week, month
    type_data = parts[2]
    
    async with state.proxy() as data:
        data["period_size"] = period_size
        data["an_type"] = f"an_{type_data}"
    
    # Показываем календарь для выбора даты
    await show_calendar(callback, state, period_size)

async def show_calendar(callback: types.CallbackQuery, state: FSMContext, period_size):
    now = datetime.now()
    async with state.proxy() as data:
        # Только если еще не установлено!

        current_month = data.get("calendar_month", now.month)
        current_year = data.get("calendar_year", now.year)

        # Если это первый запуск, сохрани в state
        data["calendar_month"] = current_month
        data["calendar_year"] = current_year

    await show_calendar_for_month(callback, state, period_size, current_month, current_year)

async def custom_period_back_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик кнопки 'Назад' в календаре"""
    async with state.proxy() as data:
        an_type = data.get("an_type", "an_1")
        type_data = an_type.split("_")[1]
    
    # Возвращаемся к выбору размера периода
    await custom_period_callback(callback, state)

async def select_date_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик выбора даты из календаря"""
    parts = callback.data.split("_")
    period_size = parts[2]  # day, week, month
    date_str = parts[3]     # YYYY-MM-DD
    
    selected_date = datetime.strptime(date_str, "%Y-%m-%d")
    
    async with state.proxy() as data:
        data["selected_date"] = selected_date
        data["period_size"] = period_size
    
    # Показываем подтверждение
    await show_date_confirmation(callback, state, selected_date, period_size)

async def show_date_confirmation(callback: types.CallbackQuery, state: FSMContext, selected_date, period_size):
    """Показывает подтверждение выбранной даты"""
    period_text = {"day": "день", "week": "неделю", "month": "месяц"}[period_size]
    
    # Рассчитываем период в зависимости от выбора
    if period_size == "day":
        start_date = selected_date
        end_date = selected_date
        period_display = selected_date.strftime("%d.%m.%Y")
    elif period_size == "week":
        # Находим понедельник недели
        start_date = selected_date - timedelta(days=selected_date.isoweekday() - 1)
        end_date = start_date + timedelta(days=6)
        period_display = f"{start_date.strftime('%d.%m')}-{end_date.strftime('%d.%m.%Y')}"
    else:  # month
        start_date = selected_date.replace(day=1)
        end_date = (start_date.replace(month=start_date.month % 12 + 1, day=1) - timedelta(days=1))
        period_display = f"{start_date.strftime('%d.%m')}-{end_date.strftime('%d.%m.%Y')}"
    
    async with state.proxy() as data:
        data["custom_start_date"] = start_date
        data["custom_end_date"] = end_date
    
    keyboard = InlineKeyboardMarkup(row_width=1)
    keyboard.add(
        InlineKeyboardButton("✅ Подтвердить", callback_data=f"confirm_custom_{period_size}"),
        InlineKeyboardButton("🔄 Выбрать другую дату", callback_data=f"custom_{period_size}_{data['an_type'].split('_')[1]}"),
        InlineKeyboardButton("🔙 Назад", callback_data="custom_period_back")
    )
    
    try:
        await callback.message.edit_text(
            f"📅 <b>Подтверждение выбора</b>\n\n"
            f"Выбранный {period_text}: <b>{period_display}</b>\n\n"
            f"Нажмите 'Подтвердить' для расчета:",
            reply_markup=keyboard
        )
    except MessageNotModified:
        # Если сообщение не изменилось, просто отвечаем на callback
        await callback.answer()
    except Exception as e:
        # Для других ошибок логируем и отвечаем на callback
        print(f"Error in show_date_confirmation: {e}")
        await callback.answer()

async def confirm_custom_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик подтверждения кастомного периода"""
    parts = callback.data.split("_")
    period_size = parts[2]  # day, week, month
    
    async with state.proxy() as data:
        start_date = data["custom_start_date"]
        end_date = data["custom_end_date"]
        an_type = data["an_type"]
    
    # Создаем фейковый callback для select_anal_period_callback
    callback.data = f"anperiod_custom_{period_size}_{an_type.split('_')[1]}"
    
    # Сохраняем кастомные даты в состоянии
    async with state.proxy() as data:
        data["custom_period"] = True
        data["custom_start_date"] = start_date
        data["custom_end_date"] = end_date
    
    # Вызываем основную функцию расчета
    await select_anal_period_callback(callback, state)    

async def calendar_navigation_callback(callback: types.CallbackQuery, state: FSMContext):
    print(f"calendar_navigation_callback triggered: {callback.data}")
    parts = callback.data.split("_")
    action = parts[0]  # prev or next
    period_size = parts[2]  # day, week, month

    data = await state.get_data()
    current_month = data.get("calendar_month", datetime.now().month)
    current_year = data.get("calendar_year", datetime.now().year)

    # ⬅️➡️ изменяем месяц
    if action == "prev":
        if current_month == 1:
            current_month = 12
            current_year -= 1
        else:
            current_month -= 1
    elif action == "next":
        if current_month == 12:
            current_month = 1
            current_year += 1
        else:
            current_month += 1

    # 💾 Сохраняем обратно в FSM
    await state.update_data(calendar_month=current_month, calendar_year=current_year)

    try:
        await show_calendar_for_month(callback, state, period_size, current_month, current_year)
    except MessageNotModified:
        await callback.answer()
    except Exception as e:
        print(f"Error in calendar_navigation_callback: {e}")
        await callback.answer()

async def show_calendar_for_month(callback: types.CallbackQuery, state: FSMContext, period_size, month, year):
    """Показывает календарь для конкретного месяца"""
    print(f"Show calendar for: {month}.{year}")
    # Создаем календарь на указанный месяц
    keyboard = InlineKeyboardMarkup(row_width=7)
    
    # Заголовок месяца
    month_date = datetime(year, month, 1)
    month_name = month_date.strftime("%B %Y")
    keyboard.add(InlineKeyboardButton(f"📅 {month_name}", callback_data="ignore"))
    
    # Дни недели
    weekdays = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    keyboard.row(*[InlineKeyboardButton(day, callback_data="ignore") for day in weekdays])
    
    # Получаем первый день месяца и количество дней
    first_day = datetime(year, month, 1)
    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)
    days_in_month = (next_month - timedelta(days=1)).day
    
    # Определяем день недели для первого дня (1 = понедельник)
    first_weekday = first_day.isoweekday()
    
    # Добавляем пустые ячейки в начале
    row = []
    for _ in range(first_weekday - 1):
        row.append(InlineKeyboardButton(" ", callback_data="ignore"))
    
    # Добавляем дни месяца
    now = datetime.now()
    for day in range(1, days_in_month + 1):
        date_str = f"{year}-{month:02d}-{day:02d}"
        callback_data = f"select_date_{period_size}_{date_str}"

        if day == now.day and month == now.month and year == now.year:
            row.append(InlineKeyboardButton(f"•{day}•", callback_data=callback_data))
        else:
            row.append(InlineKeyboardButton(str(day), callback_data=callback_data))

        if len(row) == 7:
            keyboard.row(*row)
            row = []

    # ДОБАВЛЯЕМ ОСТАВШИЕСЯ ТОЛЬКО ПОСЛЕ ЦИКЛА
    if row:
        while len(row) < 7:
            row.append(InlineKeyboardButton(" ", callback_data="ignore"))
        keyboard.row(*row)

    
    # Кнопки навигации (без ограничений по году)
    nav_row = [
        InlineKeyboardButton("◀️", callback_data=f"prev_month_{period_size}"),
        InlineKeyboardButton("🔙 Назад", callback_data="custom_period_back"),
        InlineKeyboardButton("▶️", callback_data=f"next_month_{period_size}")
    ]
    keyboard.row(*nav_row)
    
    period_text = {"day": "день", "week": "неделю", "month": "месяц"}[period_size]
    
    try:
        await callback.message.edit_text(
            f"📅 <b>Выберите {period_text} для расчета</b>\n\n"
            f"Нажмите на дату, чтобы выбрать {period_text}:",
            reply_markup=keyboard
        )
    except MessageNotModified:
        # Если сообщение не изменилось, просто отвечаем на callback
        await callback.answer()
    except Exception as e:
        # Для других ошибок логируем и отвечаем на callback
        print(f"Error in show_calendar_for_month: {e}")
        await callback.answer() 

async def ignore_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик для игнорируемых кнопок (заголовки, дни недели)"""
    await callback.answer()    

def register_analytics_handlers(dp: Dispatcher):
    dp.register_callback_query_handler(analytics_callback, text="analytics", state="*")
    dp.register_callback_query_handler(
        profitability_estimation_callback, text="profitability_estimation", state="*"
    )
    dp.register_callback_query_handler(
        top5_products_callback, text="top5_products", state="*"
    )
    dp.register_callback_query_handler(
        what_if_simulator_callback, text="what_if_simulator", state="*"
    )
    dp.register_callback_query_handler(
        product_analytics_callback, text="product_analytics", state="*"
    )
    
    dp.register_callback_query_handler(finances_handler, text="finances", state="*")
    # Пагинация и выбор артикула
    dp.register_callback_query_handler(
        handle_articles_pagination,
        lambda c: c.data in ["prev_articles_page", "next_articles_page"],
        state=AnalyticsStates.waiting_for_article,
    )
    dp.register_callback_query_handler(
        select_article_callback,
        lambda c: c.data.startswith("select_article_"),
        state=AnalyticsStates.waiting_for_article,
    )
    dp.register_callback_query_handler(
        anal_callback, lambda c: c.data.startswith("an_"), state="*"
    )

    dp.register_callback_query_handler(
    ignore_callback, 
    lambda c: c.data == "ignore", 
    state="*"
    )
    # Обработчики для календаря
    dp.register_callback_query_handler(
        calendar_navigation_callback, 
        lambda c: c.data.startswith("prev_month_") or c.data.startswith("next_month_"), 
        state="*"
    )
    
    dp.register_callback_query_handler(
        custom_period_back_callback, 
        lambda c: c.data == "custom_period_back", 
        state="*"
    )

        # Обработчик для "Выбранный период"
    dp.register_callback_query_handler(
        custom_period_callback, 
        lambda c: c.data.startswith("custom_period_"), 
        state="*"
    )

    dp.register_callback_query_handler(
        custom_period_size_callback, 
        lambda c: c.data.startswith("custom_") and not c.data.startswith("custom_period_"), 
        state="*"
    )
    
    dp.register_callback_query_handler(
        select_date_callback, 
        lambda c: c.data.startswith("select_date_"), 
        state="*"
    )
    
    dp.register_callback_query_handler(
        confirm_custom_callback, 
        lambda c: c.data.startswith("confirm_custom_"), 
        state="*"
    )
        
    dp.register_callback_query_handler(
        select_anal_period_callback, lambda c: c.data.startswith("anperiod_"), state="*"
    )
    # Ввод для симулятора
    dp.register_message_handler(
        process_price_and_cost, state=AnalyticsStates.waiting_for_price_and_cost
    )

    # Возврат в меню
    dp.register_callback_query_handler(
        back_to_analytics, text="back_to_analytics", state="*"
    )
